#views.py
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from .models import Estudiante, Profesor, Carrera, Pregunta, Respuesta, Bloque, Seleccion
from .forms import RegistrarEstudianteForm, RegistrarProfesorForm, RegistrarCarreraForm, RespuestaForm, DocumentoEstudianteForm
from django.db.models import Count, Q, F
from openpyxl import Workbook




#-------------no se usa----------------------
def index(request):
    return render(request, 'diagnostico/index.html')

#-----matematicas hijo---------------
def home(request):
    return render(request, 'diagnostico/home.html')

#-----login personalizado aunque uso el login de django---------
def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                
                if Estudiante.objects.filter(user=user).exists():
                    return redirect('pagina_estudiante')
               
                elif Profesor.objects.filter(user=user).exists() or user.is_superuser:
                    return redirect('pagina_profesor') 

        return render(request, 'registration/login.html', {'form': form })
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})
#
#
#
#--------acá el formulario que llenan los alumnos---------------
#
#
@login_required
def pagina_estudiante(request):
    # Obtener el estudiante autenticado
    estudiante = Estudiante.objects.get(user=request.user)

    # Verificar si el estudiante ya ha respondido el formulario
    if Respuesta.objects.filter(estudianteFK=estudiante).exists():
        return render(request, 'diagnostico/formulario_completado.html')
    
    # Obtener las 24 preguntas
    preguntas = Pregunta.objects.all()[:24]

    if request.method == 'POST':
        # Procesar el formulario
        for pregunta in preguntas:
            form = RespuestaForm(request.POST, prefix=f'pregunta_{pregunta.id_pregunta}')
            if form.is_valid():
                seleccion = form.cleaned_data['seleccion']
                # Tomar el primer valor de la lista (ya que solo se permite una selección)
                seleccion_unica = seleccion[0] if seleccion else ''  # Valor único o vacío
                # Guardar la respuesta en la base de datos
                Respuesta.objects.create(
                    seleccion=seleccion_unica,  # Guardar el valor único
                    preguntaFK=pregunta,
                    estudianteFK=estudiante
                )
        return redirect('respuestas_guardadas') 
    else:
        # Si no es un POST, mostrar el formulario vacío
        forms = [RespuestaForm(prefix=f'pregunta_{pregunta.id_pregunta}') for pregunta in preguntas]
        preguntas_forms = zip(preguntas, forms)

    return render(request, 'diagnostico/pagina_estudiante.html', {
        'preguntas_forms': preguntas_forms
    })

#----------------------------------------------------------


def respuestas_guardadas(request):
    return render(request, 'diagnostico/respuestas_guardadas.html')



@login_required
def pagina_profesor(request):
    return render(request, 'diagnostico/pagina_profesor.html')


#
# estos metodos solo los pueden usar los profes y son para generar nuevas instancias de profe y de estudiantes
#
#


@login_required
def registrar_estudiante(request):
    # Verificar si el usuario es un profesor
    if not (hasattr(request.user, 'profesor') or request.user.is_superuser):
        return redirect('pagina_no_autorizada')  

    if request.method == 'POST':
        form = RegistrarEstudianteForm(request.POST)
        if form.is_valid():
            # Crear usuario
            documento = form.cleaned_data['documento']
            try:
                # Verificar si ya existe un usuario con este documento
                if User.objects.filter(username=documento).exists():
                    form.add_error('documento', 'Ya existe un usuario con este documento.')
                else:
                    user = User.objects.create_user(
                        username=documento,
                        password=documento
                    )
                    # Crear estudiante
                    Estudiante.objects.create(
                        user=user,
                        documento=documento,
                        nombre=form.cleaned_data['nombre'],
                        apellido=form.cleaned_data['apellido'],
                        celular=form.cleaned_data['celular'],
                        email=form.cleaned_data['email'],
                        carrera=form.cleaned_data['carrera']
                    )
                    return redirect('registrar_estudiante') 
            except IntegrityError:
                form.add_error('documento', 'Error al crear el usuario. Intente nuevamente.')        
    else:
        form = RegistrarEstudianteForm()
    return render(request, 'diagnostico/registrar_estudiante.html', {'form': form})

@login_required
def registrar_profesor(request):
    # Verificar si el usuario es un profesor o un superUser
    if not (hasattr(request.user, 'profesor') or request.user.is_superuser):
        return redirect('pagina_no_autorizada') 

    if request.method == 'POST':
        form = RegistrarProfesorForm(request.POST)
        if form.is_valid():
            # Crear usuario
            documento = form.cleaned_data['documento']

            try:
                # Verificar si ya existe un usuario con este documento
                if User.objects.filter(username=documento).exists():
                    form.add_error('documento', 'Ya existe un usuario con este documento.')
                else:
                    user = User.objects.create_user(
                        username=documento,
                        password=documento
                    )
                    # Crear profesor
                    Profesor.objects.create(
                        user=user,
                        documento=documento,
                        nombre=form.cleaned_data['nombre'],
                        apellido=form.cleaned_data['apellido'],
                        celular=form.cleaned_data['celular'],
                        email=form.cleaned_data['email']
                    )
                    return redirect('registrar_profesor')
            except IntegrityError:
                form.add_error('documento', 'Error al crear el usuario. Intente nuevamente.')
    else:
        form = RegistrarProfesorForm()
    return render(request, 'diagnostico/registrar_profesor.html', {'form': form})

#--------no la pienso usar pero la dejo por las dudas------------
def registro_exitoso(request):
    return render(request, 'diagnostico/registro_exitoso.html')

#--------solo se usa cuando alquien intenta acceder sin credenciales-------
def pagina_no_autorizada(request):
    return render(request, 'diagnostico/pagina_no_autorizada.html')


#---------pagina principal donde están todos los enlaces a los diferentes informes----
@login_required
def informes(request):
    return render(request, 'diagnostico/allInformes.html')


#--------creo una carrera--------
@login_required
def crear_carrera(request):
    if not (hasattr(request.user, 'profesor') or request.user.is_superuser):
        return redirect('pagina_no_autorizada')  

    if request.method == 'POST':
        form = RegistrarCarreraForm(request.POST)
        if form.is_valid():
                Carrera.objects.create(
                    
                    nombre=form.cleaned_data['nombre'],
                    descripcion=form.cleaned_data['descripcion']
                    
                )
                return redirect('crear_carrera') 
              
    else:
        form = RegistrarCarreraForm()
    return render(request, 'diagnostico/crear_carrera.html', {'form': form})


#
#
#--------------------comienzan los informes-------------------------
#
#



def correctas_por_pregunta(request):
   
    # Obtener parámetros de filtro (carrera y pregunta)
    carrera_id = request.GET.get('carrera')
    pregunta_id = request.GET.get('pregunta')
    exportar_excel = request.GET.get('exportar_excel')

    # Inicializar variables
    respuestas_correctas = []
    total_correctas = None

    # Manejar el caso en que carrera_id es None o 'None'
    if carrera_id is None or carrera_id == 'None':
        carrera_id = 100  # ID de la carrera "Todas"

    # Verificar si se han enviado los filtros
    if carrera_id or pregunta_id:
        # Filtrar respuestas correctas
        respuestas = Respuesta.objects.all()

        # Filtrar por carrera (excepto si es "Todas")
        if carrera_id and carrera_id != '100':  # Asegúrate de comparar con cadena si viene como cadena
            respuestas = respuestas.filter(estudianteFK__carrera_id=carrera_id)

        # Filtrar por pregunta
        if pregunta_id:
            respuestas = respuestas.filter(preguntaFK__id_pregunta=pregunta_id)
            

        # Filtrar respuestas correctas
        respuestas_correctas = respuestas.filter(seleccion=F('preguntaFK__respuestaOK'))
        total_correctas = respuestas_correctas.count()

    # Exportar a Excel si se solicita
    if exportar_excel:
        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Respuestas Correctas"

        # Agregar encabezados
        ws.append(["Nombre del Estudiante", "Documento del Estudiante", "Pregunta Seleccionada"])

        # Agregar datos
        pregunta_seleccionada = Pregunta.objects.get(id_pregunta=pregunta_id).descripcion if pregunta_id else "N/A"
        for respuesta in respuestas_correctas:
            ws.append([
                respuesta.estudianteFK.nombre,
                respuesta.estudianteFK.documento,
                pregunta_seleccionada
            ])

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=respuestas_correctas.xlsx'
        wb.save(response)

        return response

    # Renderizar la plantilla con los datos
    return render(request, 'diagnostico/correctas_por_pregunta.html', {
        'total_correctas': total_correctas,
        'carreras': Carrera.objects.all(),
        'respuestas_correctas': respuestas_correctas,
        'preguntas': Pregunta.objects.all(),
        'carrera_seleccionada': int(carrera_id) if carrera_id else None,
        'pregunta_seleccionada': int(pregunta_id) if pregunta_id else None
    })


def opciones_por_pregunta(request):
    # Obtener parámetros de filtro (carrera y pregunta)
    carrera_id = request.GET.get('carrera')
    pregunta_id = request.GET.get('pregunta')
    exportar_excel = request.GET.get('exportar_excel')  # Obtener el parámetro exportar_excel

    # Inicializar variables
    respuestas_agrupadas = {}
    respuesta_correcta = None
    total_respuestas = 0

    # Manejar el caso en que carrera_id es None o 'None'
    if carrera_id is None or carrera_id == 'None':
        carrera_id = 100  # ID de la carrera "Todas"

    # Verificar si se han enviado los filtros
    if pregunta_id:  # Pregunta es obligatoria
        # Filtrar respuestas
        respuestas = Respuesta.objects.filter(preguntaFK_id=pregunta_id)

        # Filtrar por carrera (excepto si es "Todas")
        if carrera_id and carrera_id != '100':  # Asegúrate de comparar con cadena si viene como cadena
            respuestas = respuestas.filter(estudianteFK__carrera_id=carrera_id)

        # Agrupar respuestas por selección (A, B, C, D)
        respuestas_agrupadas = {
            'A': respuestas.filter(seleccion='A'),
            'B': respuestas.filter(seleccion='B'),
            'C': respuestas.filter(seleccion='C'),
            'D': respuestas.filter(seleccion='D'),
        }

        # Obtener la respuesta correcta de la pregunta
        respuesta_correcta = Pregunta.objects.get(id_pregunta=pregunta_id).respuestaOK

        # Calcular el total de respuestas
        total_respuestas = respuestas.count()

    # ------------Exportar a Excel------------------
    if exportar_excel:
        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Opciones por Pregunta"

        # Agregar encabezados
        ws.append(["Opción", "Cantidad de Respuestas"])

        # Agregar datos
        for opcion, respuestas in respuestas_agrupadas.items():
            ws.append([opcion, respuestas.count()])

        # Agregar la respuesta correcta y el total de respuestas
        ws.append(["Respuesta Correcta", respuesta_correcta])
        ws.append(["Total de Respuestas", total_respuestas])

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=opciones_por_pregunta.xlsx'
        wb.save(response)

        return response

    # Renderizar la plantilla con los datos
    return render(request, 'diagnostico/opciones_por_pregunta.html', {
        'respuestas_agrupadas': respuestas_agrupadas,
        'respuesta_correcta': respuesta_correcta,
        'total_respuestas': total_respuestas,
        'carreras': Carrera.objects.all(),
        'preguntas': Pregunta.objects.all(),
        'carrera_seleccionada': int(carrera_id) if carrera_id else None,
        'pregunta_seleccionada': int(pregunta_id) if pregunta_id else None,
    })

#
#
#
#
from django.db.models import F
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from .models import Respuesta, Carrera, Bloque, Estudiante

def correctas_por_bloque(request):
    # Obtener parámetros de filtro (carrera, bloque y cantidad de correctas)
    carrera_id = request.GET.get('carrera')
    bloque = request.GET.get('bloque')
    cantidad_correctas = request.GET.get('cantidad_correctas')
    exportar_excel = request.GET.get('exportar_excel')  # Obtener el parámetro exportar_excel

    # Inicializar variables
    estudiantes_filtrados = []
    bloques_disponibles = Bloque.choices  # Acceder a las opciones de Bloque directamente

    # Manejar el caso en que carrera_id es None o 'None'
    if carrera_id is None or carrera_id == 'None':
        carrera_id = 100  # ID de la carrera "Todas"

    # Filtrar estudiantes si se proporcionan los filtros
    if bloque and cantidad_correctas:
        # Convertir cantidad_correctas a entero
        cantidad_correctas = int(cantidad_correctas)

        # Filtrar estudiantes (excepto si es "Todas")
        estudiantes = Estudiante.objects.all()
        if carrera_id and carrera_id != '100':  # Asegúrate de comparar con cadena si viene como cadena
            estudiantes = estudiantes.filter(carrera_id=carrera_id)

        # Filtrar estudiantes con la cantidad de correctas en el bloque seleccionado
        for estudiante in estudiantes:
            # Contar respuestas correctas en el bloque seleccionado
            correctas = Respuesta.objects.filter(
                estudianteFK=estudiante,
                preguntaFK__block=bloque,
                seleccion=F('preguntaFK__respuestaOK')
            ).count()

            # Si coincide con la cantidad seleccionada, agregar al resultado
            if correctas == cantidad_correctas:
                estudiantes_filtrados.append({
                    'estudiante': estudiante,
                    'correctas': correctas
                })

    # ------------Exportar a Excel------------------
    if exportar_excel:
        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Correctas por Bloque"

        # Agregar encabezados
        ws.append(["Nombre del Estudiante", "Documento del Estudiante", "Cantidad de Correctas"])

        # Agregar datos
        for estudiante_info in estudiantes_filtrados:
            estudiante = estudiante_info['estudiante']
            correctas = estudiante_info['correctas']
            ws.append([estudiante.nombre, estudiante.documento, correctas])

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=correctas_por_bloque.xlsx'
        wb.save(response)

        return response

    # Renderizar la plantilla con los datos
    return render(request, 'diagnostico/correctas_por_bloque.html', {
        'estudiantes_filtrados': estudiantes_filtrados,
        'carreras': Carrera.objects.all(),
        'bloques_disponibles': bloques_disponibles,
        'carrera_seleccionada': int(carrera_id) if carrera_id else None,
        'bloque_seleccionado': bloque,
        'cantidad_correctas_seleccionada': cantidad_correctas,
    })

#
#
#
#
#


from django.db.models import F
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from .models import Respuesta, Carrera, Bloque, Estudiante

def comparacion_bloques(request):
    # Obtener parámetros de filtro (carrera, bloque1, bloque2 y cantidad de correctas)
    carrera_id = request.GET.get('carrera')
    bloque1 = request.GET.get('bloque1')
    bloque2 = request.GET.get('bloque2')
    cantidad_correctas = request.GET.get('cantidad_correctas')
    exportar_excel = request.GET.get('exportar_excel')  # Obtener el parámetro exportar_excel

    # Inicializar variables
    estudiantes_filtrados = []
    bloques_disponibles = Bloque.choices  # Acceder a las opciones de Bloque directamente

    # Manejar el caso en que carrera_id es None o 'None'
    if carrera_id is None or carrera_id == 'None':
        carrera_id = 100  # ID de la carrera "Todas"

    # Filtrar estudiantes si se proporcionan los filtros
    if bloque1 and bloque2 and cantidad_correctas:
        # Convertir cantidad_correctas a entero
        cantidad_correctas = int(cantidad_correctas)

        # Filtrar estudiantes (excepto si es "Todas")
        estudiantes = Estudiante.objects.all()
        if carrera_id and carrera_id != '100':  # Asegúrate de comparar con cadena si viene como cadena
            estudiantes = estudiantes.filter(carrera_id=carrera_id)

        # Filtrar estudiantes con la cantidad de correctas en los bloques seleccionados
        for estudiante in estudiantes:
            # Contar respuestas correctas en el primer bloque
            correctas_bloque1 = Respuesta.objects.filter(
                estudianteFK=estudiante,
                preguntaFK__block=bloque1,
                seleccion=F('preguntaFK__respuestaOK')
            ).count()

            # Contar respuestas correctas en el segundo bloque
            correctas_bloque2 = Respuesta.objects.filter(
                estudianteFK=estudiante,
                preguntaFK__block=bloque2,
                seleccion=F('preguntaFK__respuestaOK')
            ).count()

            # Si coincide con la cantidad seleccionada en ambos bloques, agregar al resultado
            if correctas_bloque1 == cantidad_correctas and correctas_bloque2 == cantidad_correctas:
                estudiantes_filtrados.append({
                    'estudiante': estudiante,
                    'correctas_bloque1': correctas_bloque1,
                    'correctas_bloque2': correctas_bloque2,
                })

    # ------------Exportar a Excel------------------
    if exportar_excel:
        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparación de Bloques"

        # Agregar encabezados
        ws.append(["Nombre del Estudiante", "Documento del Estudiante", f"Correctas en {bloque1}", f"Correctas en {bloque2}"])

        # Agregar datos
        for estudiante_info in estudiantes_filtrados:
            estudiante = estudiante_info['estudiante']
            correctas_bloque1 = estudiante_info['correctas_bloque1']
            correctas_bloque2 = estudiante_info['correctas_bloque2']
            ws.append([estudiante.nombre, estudiante.documento, correctas_bloque1, correctas_bloque2])

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=comparacion_bloques.xlsx'
        wb.save(response)

        return response

    # Renderizar la plantilla con los datos
    return render(request, 'diagnostico/comparacion_bloques.html', {
        'estudiantes_filtrados': estudiantes_filtrados,
        'carreras': Carrera.objects.all(),
        'bloques_disponibles': bloques_disponibles,
        'carrera_seleccionada': int(carrera_id) if carrera_id else None,
        'bloque1_seleccionado': bloque1,
        'bloque2_seleccionado': bloque2,
        'cantidad_correctas_seleccionada': cantidad_correctas,
    })
    
#
#
#---------busco todas las respuestas asociadadas a un documento y pueo hacer un update----
#
#


@login_required
def buscar_estudiante(request):
    if request.method == 'POST':
        form = DocumentoEstudianteForm(request.POST)
        if form.is_valid():
            documento = form.cleaned_data['documento']
            try:
                estudiante = Estudiante.objects.get(documento=documento)
                return redirect('actualizar_respuestas', estudiante_id=estudiante.user_id)
            except Estudiante.DoesNotExist:
                form.add_error('documento', 'Estudiante no encontrado')
    else:
        form = DocumentoEstudianteForm()
    
    return render(request, 'diagnostico/buscar_estudiante.html', {'form': form})



#
#
#----------actualiza respuestas, método llamado por def buscar_estudiante():---- 
#
#

@login_required
def actualizar_respuestas(request, estudiante_id):
    estudiante = get_object_or_404(Estudiante, user_id=estudiante_id)
    preguntas = Pregunta.objects.all()[:24]
    exportar_excel = request.GET.get('exportar_excel') == '1'  # Verificar si se debe exportar a Excel

    if request.method == 'POST':
        # Procesar el formulario
        for pregunta in preguntas:
            form = RespuestaForm(request.POST, prefix=f'pregunta_{pregunta.id_pregunta}')
            if form.is_valid():
                seleccion = form.cleaned_data['seleccion']
                if seleccion:  # Si hay una selección
                    seleccion_unica = seleccion[0]  # Tomar el primer valor de la lista
                    # Buscar la respuesta existente o crear una nueva
                    respuesta, created = Respuesta.objects.get_or_create(
                        preguntaFK=pregunta,
                        estudianteFK=estudiante,
                        defaults={'seleccion': seleccion_unica}
                    )
                    if not created:  # Si la respuesta ya existe, actualizarla
                        respuesta.seleccion = seleccion_unica
                        respuesta.save()
                else:
                    # Si no hay selección, eliminar la respuesta existente (si existe)
                    Respuesta.objects.filter(preguntaFK=pregunta, estudianteFK=estudiante).delete()
        return redirect('actualizar_respuestas', estudiante_id=estudiante.user_id)
    else:
        # Obtener las respuestas existentes del estudiante
        respuestas_existentes = Respuesta.objects.filter(estudianteFK=estudiante)
        respuestas_dict = {respuesta.preguntaFK.id_pregunta: respuesta.seleccion for respuesta in respuestas_existentes}

        # Crear los formularios con las respuestas existentes
        forms = []
        for pregunta in preguntas:
            seleccion_inicial = respuestas_dict.get(pregunta.id_pregunta, None)
            seleccion_inicial_lista = [seleccion_inicial] if seleccion_inicial else []  # Convertir a lista
            form = RespuestaForm(prefix=f'pregunta_{pregunta.id_pregunta}', initial={'seleccion': seleccion_inicial_lista})
            forms.append(form)

        preguntas_forms = zip(preguntas, forms)

    # Exportar a Excel si es necesario
    if exportar_excel:
        # Crear un libro de Excel y una hoja
        wb = Workbook()
        ws = wb.active
        ws.title = "Respuestas del Estudiante"

        # Agregar encabezados
        ws.append(["Pregunta", "Respuesta"])

        # Agregar datos (respuestas del estudiante)
        for respuesta in respuestas_existentes:
            ws.append([respuesta.preguntaFK.descripcion, respuesta.seleccion])

        # Crear una respuesta HTTP con el archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename=respuestas_{estudiante.documento}.xlsx'
        wb.save(response)
        
        return response

    return render(request, 'diagnostico/actualizar_respuestas.html', {
        'estudiante': estudiante,
        'preguntas_forms': preguntas_forms
    })

